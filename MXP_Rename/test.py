import os
from openpyxl import load_workbook

path = 'Y:\Python\MXP Rename\MXP_2nd_01172018_TestRun_206\\test'
HS_count = 66
Die_count = 1

if path.startswith('\\'):                          #以下这段判断语句用来处理网络地址\\192.168.1.1 或者Y:\TEst\XXx
    path_converted = '/' + path.replace("\\", "/")
else:
    path_converted = path.replace("\\", "/")               #判断结束，将路径转换成统一格式


files = os.listdir(path_converted)

if len(files) != (HS_count * Die_count + 1):
    print ("file count doesn't match!")
else:
    wb = load_workbook(path_converted + '/Gauge.xlsx')
    ws_names = wb.get_sheet_names()
    ws_gauge = wb[ws_names[0]]
    ws_die = wb[ws_names[1]]
    if (ws_gauge.max_row - 1) == HS_count and ws_gauge.max_column == 2 and (
            ws_die.max_row - 1) == Die_count and ws_die.max_column == 2:
        for file in files:
            if file.startswith('image') and file.endswith('jpg'):
                old_name = os.path.splitext(file)[0]
                point_id = int(old_name.split('-')[3])  # start with 0
                die_seq = int((int(old_name.split('-')[4]) - point_id) / HS_count)  # start with 0
                gauge_name = ws_gauge.cell(row=point_id + 2, column=2)
                die_index_X = ws_die.cell(row=die_seq + 2, column=1)
                die_index_Y = ws_die.cell(row=die_seq + 2, column=2)
                os.rename(path_converted + '/' + file,
                          path_converted + '/' + str(gauge_name.value) + '-' + '(' + str(die_index_X.value) + ',' + str(
                              die_index_Y.value) + ')' + '.jpg')
    else:
        print("Rows or columns don't match!")




