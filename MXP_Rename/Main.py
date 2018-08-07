# -*- coding: utf-8 -*-
# @Time    : 6/23/2018 17:54
# @Author  : tcgkw
# @Email   :
# @File    : Main.py
# @Software:    python3.6.5
#               pyqt5
#               pycharm 2018.1

from MXP_Rename_UI import Ui_MainWindow  # 导入uitestPyQt5.ui转换为uitestPyQt5.py中的类
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtCore import QCoreApplication
from PyQt5.QtWidgets import QFileDialog
import sys
import os
from openpyxl import load_workbook


class Mywindow(QtWidgets.QMainWindow, Ui_MainWindow):
    # 建立的是Main Window项目，故此处导入的是QMainWindow
    # class myform(QWidget,Ui_Form):如建立的是Widget项目，导入的是QWidget
    def __init__(self):
        super(Mywindow, self).__init__()
        self.setupUi(self)

#####################################主函数代码#################
    def choose_path(self):
        directory1 = QFileDialog.getExistingDirectory(self, "Choose Folder", "")  # 起始路径
        if directory1 != '':
            self.Path.setText(directory1)
            files = os.listdir(directory1)
            for file in files:
                if os.path.splitext(file)[1] == '.xlsx':
                    self.Gauge.setText(directory1 + '/' + file)
                    break
                else:
                    self.Gauge.setText('need gauge file')
        else:
            self.Gauge.setText('need gauge file')

    def choose_gauge(self):
        fname, ftype = QFileDialog.getOpenFileName(self, "Choose Gauge File", "", "Excel Files (*.xlsx);;All Files (*)")
        if fname != '':
            self.Gauge.setText(fname)
        else:
            self.Gauge.setText('need gauge file')

    def start_rename(self):
        image_path = self.Path.toPlainText()
        gauge_path = self.Gauge.toPlainText()
        hs_count = self.HS_Count.toPlainText()
        die_count = self.Die_Count.toPlainText()
        # if image_path.startswith('\\'):  # 以下这段判断语句用来处理网络地址\\192.168.1.1 或者Y:\TEst\XXx
        #     path_converted = '/' + image_path.replace("\\", "/")
        # else:
        #     path_converted = image_path.replace("\\", "/")  # 判断结束，将路径转换成统一格式
        files = os.listdir(image_path)
        for file in files:
            if os.path.splitext(file)[1] != '.bmp':
                files.remove(file)
        wb = load_workbook(gauge_path)
        ws_names = wb.get_sheet_names()
        ws_gauge = wb[ws_names[0]]
        ws_die = wb[ws_names[1]]
        if not hs_count:
            hs_count = ws_gauge.max_row - 1
        if not die_count:
            die_count = ws_die.max_row - 1

        if (ws_gauge.max_row - 1) != hs_count or ws_gauge.max_column != 2 or (
                    ws_die.max_row - 1) != die_count or ws_die.max_column != 2 or len(files) != (hs_count * die_count):
            QCoreApplication.processEvents()
            self.statusbar.showMessage("File count doesn't match! Please check image file count, gauge list and die list")
        else:
            i = 1
            for file in files:
                if file.startswith('image') and file.endswith('bmp'):
                    old_name = os.path.splitext(file)[0]
                    point_id = int(old_name.split('-')[3])  # start with 0
                    die_seq = int((int(old_name.split('-')[4]) - point_id) / hs_count)  # start with 0
                    pattern_id = ws_gauge.cell(row=point_id + 2, column=1)
                    gauge_name = ws_gauge.cell(row=point_id + 2, column=2)
                    die_index_X = ws_die.cell(row=die_seq + 2, column=1)
                    die_index_Y = ws_die.cell(row=die_seq + 2, column=2)
                    os.rename(image_path + '/' + file,
                              image_path + '/' + str(pattern_id.value) + '-' + str(gauge_name.value) + '-' + '(' + str(
                                  die_index_X.value) + ',' + str(
                                  die_index_Y.value) + ').bmp')
                    self.statusbar.showMessage('Processing.....' + str(i) + "/" + str(len(files)))
                    QCoreApplication.processEvents()
                    i += 1
            self.statusbar.showMessage('Jobs Done!' + 'image count' + hs_count * die_count)

    @pyqtSlot()
    def start_onclick(self):
        self.start_rename()

    def path_choose_onclick(self):
        self.choose_path()

    def gauge_choose_onclick(self):
        self.choose_gauge()

###########################################################


app = QtWidgets.QApplication(sys.argv)
window = Mywindow()
window.show()
# window=myform()   #如果是QWidget
#windows.show()
#app.exec_()
sys.exit(app.exec_())
