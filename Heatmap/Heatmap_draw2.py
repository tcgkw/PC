from numpy import linspace, meshgrid
from matplotlib.mlab import griddata
from matplotlib import cm
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np
from scipy.interpolate import Rbf


path = 'Y:\Python\Heat map\middle die'

wb = load_workbook(path + '\\' + 'batch.xlsx')
ws = wb.active
count = ws.max_row
list_x = []
list_y = []
list_cd = []

for i in range(count):
    print(i)
    list_x.append(ws.cell(row=i + 1, column=1).value)
    list_y.append(ws.cell(row=i + 1, column=2).value)
    list_cd.append(ws.cell(row=i + 1, column=3).value)
x = np.array(list_x, dtype=int)
y = np.array(list_y, dtype=int)
z = np.array(list_cd, dtype=float)

ti = np.linspace(-2.0, 2.0, 100)
XI, YI = np.meshgrid(ti, ti)
rbf = Rbf(x, y, z, epsilon = 2)
ZI = rbf(XI, YI)

plt.subplot(1, 1, 1)
plt.pcolor(XI, YI, ZI, cmap=cm.jet)
plt.scatter(x, y, 100, z, cmap=cm.jet)
plt.title('RBF interpolation - multiquadrics')
plt.xlim(-2, 2)
plt.ylim(-2, 2)
plt.colorbar()
