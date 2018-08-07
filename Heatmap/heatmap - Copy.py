import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np

path = 'Y:\Python\Heat map\middle die'
files = os.listdir(path)

for file in files:
    if os.path.splitext(file)[1] != '.txt':
        files.remove(file)
files.sort(key = lambda files: int(os.path.splitext(files)[0]))

# print(files)
wb = Workbook()
wb.save(path + '\\' + 'batch.xlsx')
ws = wb.active
i = 1
list2 =[]
for file in files:
    # wb = load_workbook(path + '\\' + 'batch.xlsx')
    with open(path + '\\' + file, 'r') as f: 
        str1 = f.readlines()
        for line in str1[2:]:
            temp = re.split(r'[\s(),]', line)
            list2.append(str(temp[1], temp[3], temp[5]))
            ws.cell(row=i, column=1, value=int(temp[1]) + 8192 * (int(os.path.splitext(file)[0]) % 6 ))
            ws.cell(row=i, column=2, value=int(temp[3]) + 8192 * (3 - int((int(os.path.splitext(file)[0]) / 6))))
            ws.cell(row=i, column=3, value=float(temp[5]))
            i += 1
    wb.save(path + '\\' + 'batch.xlsx')

