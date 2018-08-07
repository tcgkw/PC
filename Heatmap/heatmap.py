import os
import re
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np

path = 'Y:\Python\Heat map\edge die'
files = os.listdir(path)

x_period = 880
y_period = 880


for file in files:
    if os.path.splitext(file)[1] != '.txt':
        files.remove(file)
files.sort(key=lambda files: int(os.path.splitext(files)[0]))


wb = Workbook()
wb.save(path + '\\' + 'batch.xlsx')
ws = wb.active
i = 1
list2 = []
for file in files:
    # wb = load_workbook(path + '\\' + 'batch.xlsx')
    with open(path + '\\' + file, 'r') as f: 
        str1 = f.readlines()
        for line in str1[2:]:
            temp = re.split(r'[\s(),]', line)
            list2.append(str(int(temp[1]) + 8192 * (int(os.path.splitext(file)[0]) % 6)) + '\t' + str(int(temp[3]) + 8192 * (3 - int((int(os.path.splitext(file)[0]) / 6)))) + '\t' + temp[5])
            ws.cell(row=i, column=1, value=int(temp[1]) + 8192 * (int(os.path.splitext(file)[0]) % 6 ))
            ws.cell(row=i, column=2, value=int(temp[3]) + 8192 * (3 - int((int(os.path.splitext(file)[0]) / 6))))
            ws.cell(row=i, column=3, value=float(temp[5]))
            i += 1
    wb.save(path + '\\' + 'batch.xlsx')

# with open('C:\\Users\kgao\Documents\HMI\PyCharm\Heatmap\\Middle die data.txt', 'r') as f:
#     list1 = f.readlines()

sum = np.zeros(shape=((int(29499/y_period)+1), (int(48426/x_period))+1), dtype=float)
count = np.zeros(shape=((int(29499/y_period)+1), (int(48426/x_period))+1), dtype=int)


with open(path + '\\' + 'temp.txt', 'w') as ff:
    for line in list2:
        ff.write(line)
        ff.write('\n')


for line in list2:
    str1 = re.split(r'[\t\n]', line)
    # print(str1[0], str1[1], str1[2])
    sum[int(int(str1[1])/y_period)][int(int(str1[0])/x_period)] += float(str1[2])
    count[int(int(str1[1])/y_period)][int(int(str1[0])/x_period)] += 1

grid = sum/count

plt.imshow(grid, cmap='jet', interpolation='nearest', origin='lower', vmin=6.3, vmax=6.9)
plt.colorbar()
plt.show()
