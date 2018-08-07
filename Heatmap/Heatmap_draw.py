from numpy import linspace, meshgrid
# import matplotlib as mpl
from matplotlib.mlab import griddata
# import matplotlib as plt
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import numpy as np


path = 'Y:\Python\Heat map\middle die'


def grid(x, y, z, resX = 2400, resY = 1250):
    "Convert 3 column data to matplotlib grid"
    xi = linspace(min(x), max(x), resX)
    yi = linspace(min(y), max(y), resY)
    Z = griddata(x, y, z, xi, yi, interp='linear')
    X, Y = meshgrid(xi, yi)
    return X, Y, Z


wb = load_workbook(path + '\\' + 'batch.xlsx')
ws = wb.active
count = ws.max_row
list_x = []
list_y = []
list_cd = []

for i in range(count):
    print(i)
    list_x.append(ws.cell(row=i + 1, column=1).value)
    list_y.append(ws.cell(row=i + 1, column=2).value)
    list_cd.append(ws.cell(row=i + 1, column=3).value)
x = np.array(list_x, dtype=int)
y = np.array(list_y, dtype=int)
z = np.array(list_cd, dtype=float)

X, Y, Z = grid(x, y, z)
plt.contourf(X, Y, Z)
plt.show()
