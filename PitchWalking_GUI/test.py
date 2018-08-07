# -*- coding: utf-8 -*-
# @Time    : 6/17/2018 17:54
# @Author  : tcgkw
# @Email   :
# @File    : Main.py
# @Software:    python3.6.5
#               pyqt5
#               pycharm 2018.1

from PitchWalking_GUI import Ui_MainWindow  # 导入uitestPyQt5.ui转换为uitestPyQt5.py中的类
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSlot
import sys
import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook


class Mywindow(QtWidgets.QMainWindow, Ui_MainWindow):
    # 建立的是Main Window项目，故此处导入的是QMainWindow
    # class myform(QWidget,Ui_Form):如建立的是Widget项目，导入的是QWidget
    def __init__(self):
        super(Mywindow, self).__init__()
        self.setupUi(self)

#####################################主函数代码#################
    def start(self):
        path = self.SearchPath.toPlainText()
        direction = 0
        if self.radioButton_forX.isChecked():
            direction = 0
        elif self.radioButton_forY.isChecked():
            direction = 1
        files = os.listdir(path)
        wb = Workbook()
        wb.save(path + '\\' + 'PitchWalking_Batch.xlsx')  # PitchWalking_Batch.xlsx可以替换成自己想要的名字
        j = 0
        for file in files:
            if os.path.splitext(files[j])[1] == '.txt':  # 判断文件是否为txt格式，如果是则开始读取
                with open(path + '\\' + file, 'r') as f:
                    list1 = []
                    str1 = f.readlines()  # 读取完毕，开始按照X或Y从小到大排序
                    for line in str1[2:]:
                        string = re.split(r'[\s(),]', line)
                        list1.append([string[1], string[3], string[5]])
                    list1.sort(key=lambda list1: int(list1[direction]))  # 到此将输入的txt按照X或Y 从小到大排序完毕

                str2 = list1[0]  # 开始分段计算平均值并存到list2
                m = int(str2[direction])
                s = 0
                count_i = 0
                count_e = 0
                list2 = []
                for line in list1:
                    if (int(line[direction]) - m) < 15:
                        if line[2] != '0':
                            s += float(line[2])
                            count_i += 1
                            count_e += 1
                        else:
                            count_e += 1
                    elif count_i != 0:
                        list2.append([m, (s / count_i), count_e])
                        if line[2] != '0':
                            m = int(line[direction])
                            s = float(line[2])
                            count_i = 1
                            count_e = 1
                        else:
                            s = 0
                            count_i = 0
                            count_e = 1
                    else:
                        if line[2] != '0':
                            m = int(line[direction])
                            s = float(line[2])
                            count_i = 1
                            count_e = 1
                        else:
                            s = 0
                            count_i = 0
                            count_e = 1
                    #print('count_internal=',count_i,'count_external=',count_e,'m=', m)
                if count_i != 0:
                    list2.append([m, (s / count_i), count_e])  # 平均值计算完毕
                wb = load_workbook(path + '\\' + 'PitchWalking_Batch.xlsx')  # 开始将list2中数值存到xlsx里
                ws = wb.create_sheet(title=os.path.basename(file.split(".")[0]))
                i = 2
                if direction == 0:
                    ws.cell(row=1, column=1, value='X')
                elif direction == 1:
                    ws.cell(row=1, column=1, value='Y')
                ws.cell(row=1, column=2, value='Average')
                ws.cell(row=1, column=3, value='Count')
                ws.cell(row=1, column=4, value='5 Smoothing')
                for line in list2:
                    ws.cell(row=i, column=1, value=line[0])
                    ws.cell(row=i, column=2, value=line[1])
                    ws.cell(row=i, column=3, value=line[2])
                    i += 1
                    # list2中数值存到xlsx里完毕，下面开始计算范围为5个值的smooth
                for k in range(i - 6):
                    ws.cell(row=k + 2, column=4,
                            value='=AVERAGE(B' + str(k + 2) + ':B' + str(k + 6) + ')')  # smooth完毕，下面保存

                wb.save(path + '\\' + 'PitchWalking_Batch.xlsx')
                j += 1
            else:
                j += 1
        self.SearchPath.setText('Done!')
    @pyqtSlot()
    def on_click(self):
        self.start()

###########################################################


app = QtWidgets.QApplication(sys.argv)
window = Mywindow()
window.show()
# window=myform()   #如果是QWidget
#windows.show()
#app.exec_()
sys.exit(app.exec_())