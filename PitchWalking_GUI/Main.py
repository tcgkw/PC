# -*- coding: utf-8 -*-
# @Time    : 7/29/2018 19:54
# @Author  : tcgkw
# @Email   :
# @File    : Main.py
# @Software:    python3.7.0
#               pyqt5
#               pycharm 2018.1

from PitchWalking_GUI import Ui_MainWindow  # 导入uitestPyQt5.ui转换为uitestPyQt5.py中的类
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSlot
import sys
import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from PyQt5.QtCore import QCoreApplication
from PIL import Image
import numpy as np
from numpy.lib.stride_tricks import as_strided


class Mywindow(QtWidgets.QMainWindow, Ui_MainWindow):
    # 建立的是Main Window项目，故此处导入的是QMainWindow
    # class myform(QWidget,Ui_Form):如建立的是Widget项目，导入的是QWidget
    def __init__(self):
        super(Mywindow, self).__init__()
        self.setupUi(self)

#####################################主函数代码#################

    ######## K- means started############

    def kmeans(self, data, k):
        clusterAssment = np.zeros(shape=(data.shape[0], 2))
        clusterAssment[:,
        0] = data  # input target dataset to clusterAssment, first column is GLV value, second column is cluster index
        cluster_changed = True
        centroids = [data.max(), data.min()]

        while cluster_changed:
            cluster_changed = False

            ## for each point in data
            for i in range(data.shape[0]):
                minDist = 10000.0
                min_index = 0

                ##  find the centroid which is the closest
                for j in range(k):
                    distance = abs(data[i] - centroids[j])
                    if distance < minDist:
                        minDist = distance
                        min_index = j

                if clusterAssment[i, 1] != min_index:
                    cluster_changed = True
                    clusterAssment[i, 1] = min_index

            ## update new centroid after new cluster assigned
            for j in range(k):
                tag = np.where(clusterAssment[:, 1] == j)[0]
                if len(tag) > 0:
                    s = 0
                    for element in tag:
                        s += clusterAssment[element][0]
                    centroids[j] = s / len(tag)

        # print('Congratulations, cluster complete!')
        return centroids, clusterAssment

    ######## K- means finished ############

    # function below used to compress image for quicker calculation
    def strided_rescale(self, g, bin_fac):
        strided = as_strided(g,
                             shape=(g.shape[0] // bin_fac, g.shape[1] // bin_fac, bin_fac, bin_fac),
                             strides=((g.strides[0] * bin_fac, g.strides[1] * bin_fac) + g.strides))
        return strided.mean(axis=-1).mean(axis=-1)

    # function above used to compress image for quicker calculation

    # function below is to detect edge and return edge type and edge pixel number

    def detect_edge(self, path):
        original_img = np.array(Image.open(path).convert('L'), 'i')  # open target image
        ratio = int(original_img.shape[0] / 512)
        # ratio can be defined by user, now all images will be compressed to 512X512 px
        rescaled_image = np.flipud(np.uint8(self.strided_rescale(original_img, ratio)))

        # img = Image.fromarray(rescaled_image, 'L')  # transfer rescaled array to bmp for check purpose
        # img.save('E:\Python\Image boundary\\img.bmp')  # save rescaled image to bmp for check purpose

        # Start to calculate  profile of image
        sum_x = np.zeros(shape=(rescaled_image.shape[1]), dtype=int)
        sum_y = np.zeros(shape=(rescaled_image.shape[0]), dtype=int)

        for c in range(rescaled_image.shape[0]):
            for r in range(rescaled_image.shape[1]):
                sum_x[c] += rescaled_image[r][c]

        for r in range(rescaled_image.shape[1]):
            for c in range(rescaled_image.shape[0]):
                sum_y[r] += rescaled_image[r][c]

        x_profile = sum_x / rescaled_image.shape[0]
        y_profile = sum_y / rescaled_image.shape[1]
        # Finish to calculate both X and Y profile of image

        # Start to do K-means
        centroids, cluster1 = self.kmeans(x_profile, 2)
        centroids, cluster2 = self.kmeans(y_profile, 2)
        # K-means done

        cluster_x = cluster1[:, 1]
        cluster_y = cluster2[:, 1]

        x_diff_abs = np.absolute(np.diff(x_profile))
        y_diff_abs = np.absolute(np.diff(y_profile))

        x_peak = np.where(x_diff_abs > (x_diff_abs.mean() * 3))[0]  # 4 times larger than  mean is defined as a peak
        y_peak = np.where(y_diff_abs > (y_diff_abs.mean() * 3))[0]
        # find the edge of pattern by the  diff peak which  is closest to cluster with higher GLV
        edge_px_x = 0
        edge_px_y = 0
        # detect x direction edge and get edge px number
        if x_peak.size > 0:
            x_l = x_peak[0]
            x_r = x_peak[-1]
            if np.mean(cluster_x[:int(len(cluster_x) / 2)]) < np.mean(cluster_x[int(len(cluster_x) / 2):]):
                x_dir = 'right edge'
                edge_px_x = x_l * ratio
            elif np.mean(cluster_x[:int(len(cluster_x) / 2)]) > np.mean(cluster_x[int(len(cluster_x) / 2):]):
                x_dir = 'left edge'
                edge_px_x = x_r * ratio
            else:
                x_dir = 'X direction edge detection error!'
        else:
            x_dir = 'center'
        # detect x direction edge
        # detect y direction edge
        if y_peak.size > 0:
            y_b = y_peak[0]
            y_t = y_peak[-1]
            if np.mean(cluster_y[:int(len(cluster_y) / 2)]) < np.mean(cluster_y[int(len(cluster_y) / 2):]):
                y_dir = 'top edge'
                edge_px_y = y_b * ratio
            elif np.mean(cluster_y[:int(len(cluster_y) / 2)]) > np.mean(cluster_y[int(len(cluster_y) / 2):]):
                y_dir = 'bottom edge'
                edge_px_y = y_t * ratio
            else:
                y_dir = 'Y direction edge detection error!'
        else:
            y_dir = 'center'
        # detect y direction edge

        # plt.plot(x_diff_abs)
        # plt.xlabel('X_profile')
        # plt.show()
        # plt.plot(y_diff_abs)
        # plt.xlabel('Y_profile')
        # plt.show()
        return {'x_dir': x_dir, 'edge_px_x': edge_px_x, 'y_dir': y_dir, 'edge_px_y': edge_px_y}

    #####  trim_list function is used to delete unwanted CD value acoording to image with the same name ####
    def trim_list(self, path, file, imported_list, list_bmp, b):  # b is buffer range to further reduce measure area
        output_list = []
        if os.path.splitext(file)[0] in list_bmp:
            result = self.detect_edge(path + '\\' + os.path.splitext(file)[0] + '.bmp')
            if result['x_dir'] == 'left edge' and result['y_dir'] == 'top edge':
                start_left = result['edge_px_x']
                end_top = result['edge_px_y']
                start_bottom = min(int(line[1]) for line in imported_list)
                end_right = max(int(line[0]) for line in imported_list)
            elif result['x_dir'] == 'center' and result['y_dir'] == 'top edge':
                start_left = min(int(line[0]) for line in imported_list)
                end_top = result['edge_px_y']
                start_bottom = min(int(line[1]) for line in imported_list)
                end_right = max(int(line[0]) for line in imported_list)
            elif result['x_dir'] == 'right edge' and result['y_dir'] == 'top edge':
                start_left = min(int(line[0]) for line in imported_list)
                end_top = result['edge_px_y']
                start_bottom = min(int(line[1]) for line in imported_list)
                end_right = result['edge_px_x']
            elif result['x_dir'] == 'left edge' and result['y_dir'] == 'center':
                start_left = result['edge_px_x']
                end_top = max(int(line[1]) for line in imported_list)
                start_bottom = min(int(line[1]) for line in imported_list)
                end_right = max(int(line[0]) for line in imported_list)
            elif result['x_dir'] == 'center' and result['y_dir'] == 'center':
                start_left = min(int(line[0]) for line in imported_list)
                end_top = max(int(line[1]) for line in imported_list)
                start_bottom = min(int(line[1]) for line in imported_list)
                end_right = max(int(line[0]) for line in imported_list)
            elif result['x_dir'] == 'right edge' and result['y_dir'] == 'center':
                start_left = min(int(line[0]) for line in imported_list)
                end_top = max(int(line[1]) for line in imported_list)
                start_bottom = min(int(line[1]) for line in imported_list)
                end_right = result['edge_px_x']
            elif result['x_dir'] == 'left edge' and result['y_dir'] == 'bottom edge':
                start_left = result['edge_px_x']
                end_top = max(int(line[1]) for line in imported_list)
                start_bottom = result['edge_px_y']
                end_right = max(int(line[0]) for line in imported_list)
            elif result['x_dir'] == 'center' and result['y_dir'] == 'bottom edge':
                start_left = min(int(line[0]) for line in imported_list)
                end_top = max(int(line[1]) for line in imported_list)
                start_bottom = result['edge_px_y']
                end_right = max(int(line[0]) for line in imported_list)
            elif result['x_dir'] == 'right edge' and result['y_dir'] == 'bottom edge':
                start_left = min(int(line[0]) for line in imported_list)
                end_top = max(int(line[1]) for line in imported_list)
                start_bottom = result['edge_px_y']
                end_right = result['edge_px_x']
            else:
                self.statusbar.showMessage('Edge Detection Error!')

            for line in imported_list:
                if int(line[0]) < (start_left + b) or int(line[0]) > (end_right - b) or int(
                        line[1]) < (start_bottom + b) or int(line[1]) > (end_top - b):
                    continue
                else:
                    output_list.append(line)
        return output_list

    def start(self):
        path = self.SearchPath.toPlainText()
        direction = 0
        if self.shrink_px.toPlainText() != '':
            b = int(self.shrink_px.toPlainText())
        else:
            b = 0
        if self.customized_range.toPlainText() != '':
            customized_range = int(self.customized_range.toPlainText())
        else:
            customized_range = 10
        if self.upper_limit.toPlainText() != '':
            u_limit = float(self.upper_limit.toPlainText())
        else:
            u_limit = 999
        if self.lower_limit.toPlainText() != '':
            l_limit = float(self.lower_limit.toPlainText())
        else:
            l_limit = 0
        if self.radioButton_forX.isChecked():
            direction = 0
        elif self.radioButton_forY.isChecked():
            direction = 1
        files = os.listdir(path)
        wb = Workbook()
        wb.save(path + '\\' + 'PitchWalking_Batch.xlsx')  # PitchWalking_Batch.xlsx可以替换成自己想要的名字
        j = 0
        list_bmp = []
        for file in files:
            if os.path.splitext(file)[1] == '.bmp':  # 判断文件是否为bmp格式，如果是则存入list_bmp 备用
                list_bmp.append(os.path.splitext(file)[0])

        for file in files:
            if os.path.splitext(files[j])[1] == '.txt':  # 判断文件是否为txt格式，如果是则开始读取
                self.statusbar.showMessage('Processing.....' + file)
                QCoreApplication.processEvents()
                with open(path + '\\' + file, 'r') as f:
                    list1 = []
                    str1 = f.readlines()  # 读取完毕，开始按照X或Y从小到大排序
                    for line in str1[2:]:
                        string = re.split(r'[\s(),]', line)
                        list1.append([string[1], string[3], string[5]])

                    # print(len(list1))

                    list1 = self.trim_list(path, file, list1, list_bmp, b)       ### 对list1进行trim，如果同一folder下有同名bmp，则会进行edge detection 并删除cell区以外的cd值
                    # print(len(list1))
                    # with open('E:\Python\Image boundary\\temp.txt', 'w') as w:
                    #     for line in list1:
                    #         w.write(line)
                    list1.sort(key=lambda list1: int(list1[direction]))  # 到此将输入的txt按照X或Y 从小到大排序完毕
                    # print((list1[0]))

                str2 = list1[0]  # 开始分段计算平均值并存到list2
                m = int(str2[direction])
                s = 0
                count_i = 0
                count_e = 0
                list2 = []
                for line in list1:
                    if (int(line[direction]) - m) < customized_range:
                        if l_limit < float(line[2]) < u_limit:
                            s += float(line[2])
                            count_i += 1
                            count_e += 1
                        else:
                            count_e += 1
                    elif count_i != 0:
                        list2.append([m, (s / count_i), count_e])
                        if l_limit < float(line[2]) < u_limit:
                            m = int(line[direction])
                            s = float(line[2])
                            count_i = 1
                            count_e = 1
                        else:
                            s = 0
                            count_i = 0
                            count_e = 1
                    else:
                        if l_limit < float(line[2]) < u_limit:
                            m = int(line[direction])
                            s = float(line[2])
                            count_i = 1
                            count_e = 1
                        else:
                            s = 0
                            count_i = 0
                            count_e = 1
                if count_i != 0:
                    list2.append([m, (s / count_i), count_e])  # 平均值计算完毕

                wb = load_workbook(path + '\\' + 'PitchWalking_Batch.xlsx')  # 开始将list2中数值存到xlsx里
                ws = wb.create_sheet(title=os.path.basename(file.split(".")[0]))
                i = 2
                ws.cell(row=1, column=2, value='Average')
                ws.cell(row=1, column=3, value='Count')
                ws.cell(row=1, column=4, value='5 Smoothing')
                if direction == 0:
                    ws.cell(row=1, column=1, value='X')
                elif direction == 1:
                    ws.cell(row=1, column=1, value='Y')
                for line in list2:
                    ws.cell(row=i, column=1, value=line[0])
                    ws.cell(row=i, column=2, value=line[1])
                    ws.cell(row=i, column=3, value=line[2])
                    i += 1
                    # list2中数值存到xlsx里完毕，下面开始计算范围为5个值的smooth
                for k in range(i - 6):
                    ws.cell(row=k + 2, column=4,
                            value='=AVERAGE(B' + str(k + 2) + ':B' + str(k + 6) + ')')  # smooth完毕，下面保存

                wb.save(path + '\\' + 'PitchWalking_Batch.xlsx')
                j += 1
            else:
                j += 1
        # self.SearchPath.setText('Done!')
        self.statusbar.showMessage('Jobs Done!')



    @pyqtSlot()
    def on_click(self):
        self.start()

###########################################################


app = QtWidgets.QApplication(sys.argv)
window = Mywindow()
window.show()
# window=myform()   #如果是QWidget
#windows.show()
#app.exec_()
sys.exit(app.exec_())
